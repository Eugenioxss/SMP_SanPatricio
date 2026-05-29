from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
EXTRACTED_DIR = BASE_DIR / "extracted_tables"
DISCREPANCY_REPORT_PATH = BASE_DIR / "discrepancies" / "discrepancy_report.txt"
OUTPUT_PATH = BASE_DIR / "San_Patricio_Dashboard_Marzo_2025.xlsx"

MONTHLY_BALANCE_FILES = {
    "febrero": EXTRACTED_DIR / "Balanza_Febrero__tabla_1.csv",
    "marzo": EXTRACTED_DIR / "Balanza_Marzo__tabla_1.csv",
}

OPERATIVE_FILES = {
    "levels": EXTRACTED_DIR / "Operativo_Marzo__tabla_1.csv",
    "summary": EXTRACTED_DIR / "Operativo_Marzo__tabla_2.csv",
    "aging": EXTRACTED_DIR / "Operativo_Marzo__tabla_3.csv",
    "students": EXTRACTED_DIR / "Operativo_Marzo__tabla_4.csv",
    "operational_summary": EXTRACTED_DIR / "Operativo_Marzo__tabla_5.csv",
}


DARK_BLUE = "0F2D52"
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
SOFT_BLUE = "EAF3FA"
LIGHT_GRAY = "F3F6FA"
WHITE = "FFFFFF"
GREEN = "1F7A1F"
RED = "C00000"
ORANGE = "C55A11"


TITLE_FILL = PatternFill("solid", fgColor=DARK_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=BLUE)
CARD_FILL = PatternFill("solid", fgColor=SOFT_BLUE)
SUBTLE_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
WHITE_FONT = Font(color=WHITE, bold=True)
TITLE_FONT = Font(color=WHITE, bold=True, size=14)
SECTION_FONT = Font(color=WHITE, bold=True, size=11)
LABEL_FONT = Font(bold=True, color=DARK_BLUE)
BODY_FONT = Font(color="222222")
BOLD_FONT = Font(bold=True, color="222222")
SMALL_FONT = Font(size=10, color="444444")
MUTED_FONT = Font(size=9, italic=True, color="666666")

THIN_SIDE = Side(style="thin", color="D0D7E2")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CURRENCY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.0%'
NUMBER_FORMAT = '#,##0'
TEXT_WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top")


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Missing required input file: {path}")
    return pd.read_csv(path)


def _normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _get_column(df: pd.DataFrame, column_name: str) -> str:
    normalized = {_normalize_text(column): column for column in df.columns}
    target = normalized.get(_normalize_text(column_name))
    if target is None:
        raise KeyError(f"Column not found: {column_name}. Available: {list(df.columns)}")
    return target


def _to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _load_balance_frames() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    febrero = _read_csv(MONTHLY_BALANCE_FILES["febrero"])
    marzo = _read_csv(MONTHLY_BALANCE_FILES["marzo"])

    feb = febrero.loc[:, [_get_column(febrero, "Cuenta"), _get_column(febrero, "Descripción"), _get_column(febrero, "Saldo final")]].rename(
        columns={
            _get_column(febrero, "Cuenta"): "Cuenta",
            _get_column(febrero, "Descripción"): "Descripción Febrero",
            _get_column(febrero, "Saldo final"): "Saldo Febrero",
        }
    )
    mar = marzo.loc[:, [_get_column(marzo, "Cuenta"), _get_column(marzo, "Descripción"), _get_column(marzo, "Saldo final")]].rename(
        columns={
            _get_column(marzo, "Cuenta"): "Cuenta",
            _get_column(marzo, "Descripción"): "Descripción Marzo",
            _get_column(marzo, "Saldo final"): "Saldo Marzo",
        }
    )

    feb["Saldo Febrero"] = _to_numeric(feb["Saldo Febrero"])
    mar["Saldo Marzo"] = _to_numeric(mar["Saldo Marzo"])

    merged = feb.merge(mar, on="Cuenta", how="outer")
    merged["Descripción"] = merged["Descripción Marzo"].combine_first(merged["Descripción Febrero"])
    merged["Saldo Febrero"] = merged["Saldo Febrero"].fillna(0)
    merged["Saldo Marzo"] = merged["Saldo Marzo"].fillna(0)
    merged["Mensual Marzo"] = merged["Saldo Marzo"] - merged["Saldo Febrero"]

    return febrero, marzo, merged


def _financial_kpis(merged: pd.DataFrame) -> dict[str, float]:
    account_series = merged["Cuenta"].astype(str)
    saldo_febrero = merged["Saldo Febrero"]
    saldo_marzo = merged["Saldo Marzo"]
    mensual_marzo = merged["Mensual Marzo"]

    revenue_mask = account_series.str.startswith("4")
    expense_mask = account_series.str.startswith("5")
    payroll_mask = account_series.str.startswith(("5110", "5120", "5130", "5210", "5220", "5230"))
    depreciation_mask = account_series.str.startswith("5900")

    revenue_monthly = -mensual_marzo.loc[revenue_mask].sum(min_count=1)
    expenses_monthly = mensual_marzo.loc[expense_mask].sum(min_count=1)
    depreciation_monthly = mensual_marzo.loc[depreciation_mask].sum(min_count=1)
    expenses_monthly_ex_depr = expenses_monthly - depreciation_monthly
    payroll_monthly = mensual_marzo.loc[payroll_mask].sum(min_count=1)
    ebitda_monthly = revenue_monthly - expenses_monthly_ex_depr
    margin_monthly = ebitda_monthly / revenue_monthly if revenue_monthly else 0
    payroll_pct_monthly = payroll_monthly / revenue_monthly if revenue_monthly else 0

    revenue_ytd = -saldo_marzo.loc[revenue_mask].sum(min_count=1)
    expenses_ytd = saldo_marzo.loc[expense_mask].sum(min_count=1)
    depreciation_ytd = saldo_marzo.loc[depreciation_mask].sum(min_count=1)
    expenses_ytd_ex_depr = expenses_ytd - depreciation_ytd
    payroll_ytd = saldo_marzo.loc[payroll_mask].sum(min_count=1)
    ebitda_ytd = revenue_ytd - expenses_ytd_ex_depr
    margin_ytd = ebitda_ytd / revenue_ytd if revenue_ytd else 0
    payroll_pct_ytd = payroll_ytd / revenue_ytd if revenue_ytd else 0

    return {
        "revenue_monthly": float(revenue_monthly),
        "expenses_monthly": float(expenses_monthly),
        "expenses_monthly_ex_depr": float(expenses_monthly_ex_depr),
        "payroll_monthly": float(payroll_monthly),
        "ebitda_monthly": float(ebitda_monthly),
        "margin_monthly": float(margin_monthly),
        "payroll_pct_monthly": float(payroll_pct_monthly),
        "revenue_ytd": float(revenue_ytd),
        "expenses_ytd": float(expenses_ytd),
        "expenses_ytd_ex_depr": float(expenses_ytd_ex_depr),
        "payroll_ytd": float(payroll_ytd),
        "ebitda_ytd": float(ebitda_ytd),
        "margin_ytd": float(margin_ytd),
        "payroll_pct_ytd": float(payroll_pct_ytd),
    }


def _operational_kpis() -> dict[str, object]:
    levels = _read_csv(OPERATIVE_FILES["levels"])
    summary = _read_csv(OPERATIVE_FILES["summary"])
    aging = _read_csv(OPERATIVE_FILES["aging"])
    students = _read_csv(OPERATIVE_FILES["students"])

    nivel_col_levels = _get_column(levels, "Nivel")
    alumnos_col_levels = _get_column(levels, "Alumnos activos")
    total_levels_mask = _text_mask(levels[nivel_col_levels], r"^total$")
    detail_levels_mask = ~total_levels_mask
    active_students_levels = _to_numeric(levels.loc[detail_levels_mask, alumnos_col_levels]).sum(min_count=1)
    if pd.isna(active_students_levels):
        active_students_levels = _to_numeric(levels.loc[total_levels_mask, alumnos_col_levels]).sum(min_count=1)

    nivel_col_summary = _get_column(summary, "Nivel")
    alumnos_col_summary = _get_column(summary, "Alumnos")
    facturado_col_summary = _get_column(summary, "Facturado del mes")
    cobrado_col_summary = _get_column(summary, "Cobrado del mes")
    total_summary_mask = _text_mask(summary[nivel_col_summary], r"^total$")
    detail_summary_mask = ~total_summary_mask

    total_students_summary = _to_numeric(summary.loc[detail_summary_mask, alumnos_col_summary]).sum(min_count=1)
    total_facturado = _to_numeric(summary.loc[detail_summary_mask, facturado_col_summary]).sum(min_count=1)
    total_cobrado = _to_numeric(summary.loc[detail_summary_mask, cobrado_col_summary]).sum(min_count=1)

    if pd.isna(total_students_summary):
        total_students_summary = _to_numeric(summary.loc[total_summary_mask, alumnos_col_summary]).sum(min_count=1)
    if pd.isna(total_facturado):
        total_facturado = _to_numeric(summary.loc[total_summary_mask, facturado_col_summary]).sum(min_count=1)
    if pd.isna(total_cobrado):
        total_cobrado = _to_numeric(summary.loc[total_summary_mask, cobrado_col_summary]).sum(min_count=1)

    total_students_summary = 0 if pd.isna(total_students_summary) else total_students_summary
    total_facturado = 0 if pd.isna(total_facturado) else total_facturado
    total_cobrado = 0 if pd.isna(total_cobrado) else total_cobrado

    cobranza_pct = total_cobrado / total_facturado if total_facturado else 0
    ticket_promedio = total_facturado / total_students_summary if total_students_summary else 0

    aging_label_col = _get_column(aging, "Antigüedad")
    saldo_marzo_col = _get_column(aging, "Saldo marzo")
    total_aging_mask = _text_mask(aging[aging_label_col], r"^total$")
    current_aging_mask = _text_mask(aging[aging_label_col], r"^al corriente$")
    overdue_mask = ~(total_aging_mask | current_aging_mask)

    aging_overdue = aging.loc[overdue_mask, [aging_label_col, saldo_marzo_col]].copy()
    aging_overdue[saldo_marzo_col] = _to_numeric(aging_overdue[saldo_marzo_col]).fillna(0)
    cartera_vencida_total = aging_overdue[saldo_marzo_col].sum(min_count=1)

    cartera_total = _to_numeric(aging.loc[total_aging_mask, saldo_marzo_col]).sum(min_count=1)
    if pd.isna(cartera_total) or cartera_total == 0:
        cartera_total = _to_numeric(aging.loc[~total_aging_mask, saldo_marzo_col]).sum(min_count=1)

    cartera_vencida_total = 0 if pd.isna(cartera_vencida_total) else cartera_vencida_total
    cartera_total = 0 if pd.isna(cartera_total) else cartera_total
    cartera_vencida_pct = cartera_vencida_total / cartera_total if cartera_total else 0

    return {
        "active_students_levels": float(active_students_levels),
        "total_students_summary": float(total_students_summary),
        "total_facturado": float(total_facturado),
        "total_cobrado": float(total_cobrado),
        "cobranza_pct": float(cobranza_pct),
        "ticket_promedio": float(ticket_promedio),
        "cartera_vencida_total": float(cartera_vencida_total),
        "cartera_total": float(cartera_total),
        "cartera_vencida_pct": float(cartera_vencida_pct),
        "aging_overdue": aging_overdue,
        "students": students,
        "summary": summary,
        "levels": levels,
    }


def _text_mask(series: pd.Series, pattern: str) -> pd.Series:
    return series.astype(str).str.contains(pattern, case=False, na=False)


def _format_money(value: float) -> str:
    return f"${value:,.2f}"


def _write_title(ws, title: str, max_col: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(1, 1)
    cell.value = title
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = CENTER
    for col in range(1, max_col + 1):
        ws.cell(1, col).fill = TITLE_FILL


def _write_methodology_sheet(ws) -> None:
    _write_title(ws, "Metodología y Diseño", 8)
    ws.merge_cells("A3:H6")
    ws["A3"] = (
        "Diseño del Dashboard: Se optó por una arquitectura ETL en Python. En lugar de un workbook de Excel con "
        "fórmulas manuales propensas a errores, este documento es el output automatizado de un script. Para reutilizarlo "
        "en el cierre de abril, simplemente se procesa el nuevo archivo fuente en el pipeline, garantizando integridad de "
        "datos y escalabilidad para todo el portafolio de colegios."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"].font = BODY_FONT
    ws["A3"].fill = CARD_FILL

    ws["A8"] = "Arquitectura"
    ws["A8"].font = LABEL_FONT
    ws["A9"] = "1. Extracción de CSVs desde el workbook fuente."
    ws["A10"] = "2. Cálculo vectorizado de KPIs financieros y operativos."
    ws["A11"] = "3. Publicación de un dashboard ejecutivo en Excel."
    for row in range(9, 12):
        ws[f"A{row}"].font = BODY_FONT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.row_dimensions[3].height = 78


def _write_reconciliation_sheet(ws, report_text: str) -> None:
    _write_title(ws, "Reconciliación", 4)
    lines = report_text.splitlines()
    row = 3
    for line in lines:
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = LEFT
        cell.border = Border(bottom=THIN_SIDE)
        if line.strip().endswith(":") and not line.startswith("  "):
            cell.font = Font(bold=True, color=DARK_BLUE)
            cell.fill = SUBTLE_FILL
        elif line.strip().endswith(":"):
            cell.font = BOLD_FONT
        else:
            cell.font = BODY_FONT
        row += 1

    ws.column_dimensions["A"].width = 120
    for i in range(3, row + 2):
        ws.row_dimensions[i].height = 20
    ws.freeze_panes = "A3"


def _apply_table_header(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_value_cell(cell, value_type: str) -> None:
    cell.border = THIN_BORDER
    if value_type == "currency":
        cell.number_format = CURRENCY_FORMAT
    elif value_type == "percent":
        cell.number_format = PERCENT_FORMAT
    elif value_type == "number":
        cell.number_format = NUMBER_FORMAT
    cell.alignment = RIGHT
    cell.font = BODY_FONT


def _style_label_cell(cell) -> None:
    cell.border = THIN_BORDER
    cell.alignment = LEFT
    cell.font = LABEL_FONT
    cell.fill = CARD_FILL


def _write_dashboard_sheet(ws, financial: dict[str, float], operational: dict[str, object]) -> None:
    _write_title(ws, "San Patricio Dashboard - Marzo 2025", 10)
    ws.freeze_panes = "A3"

    # Section headers
    ws["A3"] = "KPIs Financieros"
    _apply_table_header(ws, 3, 1, 3)

    ws["E3"] = "KPIs Operativos"
    _apply_table_header(ws, 3, 5, 7)

    ws["A13"] = "Cartera Vencida"
    _apply_table_header(ws, 13, 1, 3)

    ws["E13"] = "Balance Mensual vs YTD"
    _apply_table_header(ws, 13, 5, 7)

    # Financial KPIs table
    financial_rows = [
        ("Revenue", financial["revenue_monthly"], financial["revenue_ytd"], "currency"),
        ("Gastos / Costos", financial["expenses_monthly_ex_depr"], financial["expenses_ytd_ex_depr"], "currency"),
        ("Nómina", financial["payroll_monthly"], financial["payroll_ytd"], "currency"),
        ("EBITDA", financial["ebitda_monthly"], financial["ebitda_ytd"], "currency"),
        ("Margen EBITDA", financial["margin_monthly"], financial["margin_ytd"], "percent"),
        ("Nómina % Revenue", financial["payroll_pct_monthly"], financial["payroll_pct_ytd"], "percent"),
    ]
    ws["A4"] = "Métrica"
    ws["B4"] = "Mensual"
    ws["C4"] = "YTD"
    _apply_table_header(ws, 4, 1, 3)

    for idx, (label, monthly_value, ytd_value, value_type) in enumerate(financial_rows, start=5):
        _style_label_cell(ws.cell(row=idx, column=1, value=label))
        monthly_cell = ws.cell(row=idx, column=2, value=monthly_value)
        ytd_cell = ws.cell(row=idx, column=3, value=ytd_value)
        _style_value_cell(monthly_cell, value_type)
        _style_value_cell(ytd_cell, value_type)

    # Operational KPIs table
    ws["E4"] = "Métrica"
    ws["F4"] = "Valor"
    ws["G4"] = "Detalle"
    _apply_table_header(ws, 4, 5, 7)

    operational_rows = [
        ("Alumnos activos", operational["total_students_summary"], "TOTAL alumnos en la tabla de niveles", "number"),
        ("Cobranza / Facturación", operational["cobranza_pct"], f"{_format_money(operational['total_cobrado'])} / {_format_money(operational['total_facturado'])}", "percent"),
        ("Ticket promedio", operational["ticket_promedio"], "Facturado / alumnos activos", "currency"),
        ("Cartera vencida", operational["cartera_vencida_total"], "Suma de buckets vencidos", "currency"),
        ("Cartera vencida %", operational["cartera_vencida_pct"], "Vencida / saldo total", "percent"),
    ]

    for idx, (label, value, detail, value_type) in enumerate(operational_rows, start=5):
        _style_label_cell(ws.cell(row=idx, column=5, value=label))
        value_cell = ws.cell(row=idx, column=6, value=value)
        _style_value_cell(value_cell, value_type)
        detail_cell = ws.cell(row=idx, column=7, value=detail)
        detail_cell.border = THIN_BORDER
        detail_cell.alignment = LEFT
        detail_cell.font = BODY_FONT

    # Cartera breakdown table
    ws["A14"] = "Bucket"
    ws["B14"] = "Saldo marzo"
    ws["C14"] = "% del total"
    _apply_table_header(ws, 14, 1, 3)

    overdue = operational["aging_overdue"].copy()
    overdue_labels_col = overdue.columns[0]
    overdue_value_col = overdue.columns[1]
    overdue["Pct"] = overdue[overdue_value_col] / operational["cartera_total"] if operational["cartera_total"] else 0

    for idx, (_, row) in enumerate(overdue.iterrows(), start=15):
        bucket_name = row[overdue_labels_col]
        amount = row[overdue_value_col]
        pct = row["Pct"]
        _style_label_cell(ws.cell(row=idx, column=1, value=bucket_name))
        _style_value_cell(ws.cell(row=idx, column=2, value=float(amount)), "currency")
        _style_value_cell(ws.cell(row=idx, column=3, value=float(pct)), "percent")

    total_row = 15 + len(overdue)
    ws.cell(row=total_row, column=1, value="TOTAL CARTERA VENCIDA")
    _style_label_cell(ws.cell(row=total_row, column=1))
    _style_value_cell(ws.cell(row=total_row, column=2, value=operational["cartera_vencida_total"]), "currency")
    _style_value_cell(ws.cell(row=total_row, column=3, value=operational["cartera_vencida_pct"]), "percent")

    # Balance comparison summary for chart data
    ws["E14"] = "Indicador"
    ws["F14"] = "Mensual"
    ws["G14"] = "YTD"
    _apply_table_header(ws, 14, 5, 7)

    comparison_rows = [
        ("Revenue", financial["revenue_monthly"], financial["revenue_ytd"], "currency"),
        ("EBITDA", financial["ebitda_monthly"], financial["ebitda_ytd"], "currency"),
        ("Nómina", financial["payroll_monthly"], financial["payroll_ytd"], "currency"),
        ("Gastos / Costos", financial["expenses_monthly_ex_depr"], financial["expenses_ytd_ex_depr"], "currency"),
    ]
    for idx, (label, monthly_value, ytd_value, value_type) in enumerate(comparison_rows, start=15):
        _style_label_cell(ws.cell(row=idx, column=5, value=label))
        _style_value_cell(ws.cell(row=idx, column=6, value=monthly_value), value_type)
        _style_value_cell(ws.cell(row=idx, column=7, value=ytd_value), value_type)

    # Hidden data for charts
    chart_start = 2
    chart_col = 10  # J
    ws.cell(row=chart_start, column=chart_col, value="Indicador")
    ws.cell(row=chart_start, column=chart_col + 1, value="Mensual")
    ws.cell(row=chart_start, column=chart_col + 2, value="YTD")
    for offset, (label, monthly_value, ytd_value, _) in enumerate(comparison_rows, start=1):
        ws.cell(row=chart_start + offset, column=chart_col, value=label)
        ws.cell(row=chart_start + offset, column=chart_col + 1, value=monthly_value)
        ws.cell(row=chart_start + offset, column=chart_col + 2, value=ytd_value)

    pie_start = 10
    ws.cell(row=pie_start, column=chart_col, value="Bucket")
    ws.cell(row=pie_start, column=chart_col + 1, value="Saldo")
    for offset, (_, row) in enumerate(overdue.iterrows(), start=1):
        ws.cell(row=pie_start + offset, column=chart_col, value=row[overdue_labels_col])
        ws.cell(row=pie_start + offset, column=chart_col + 1, value=float(row[overdue_value_col]))

    # Charts
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "KPIs Financieros: Mensual vs YTD"
    bar_chart.y_axis.title = "MXN"
    bar_chart.x_axis.title = "Indicador"
    data = Reference(ws, min_col=11, max_col=12, min_row=2, max_row=6)
    cats = Reference(ws, min_col=10, min_row=3, max_row=6)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.height = 7.5
    bar_chart.width = 13
    ws.add_chart(bar_chart, "I3")

    pie_chart = PieChart()
    pie_chart.title = "Cartera Vencida por Bucket"
    data = Reference(ws, min_col=11, min_row=11, max_row=14)
    labels = Reference(ws, min_col=10, min_row=11, max_row=14)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.height = 7.0
    pie_chart.width = 10.5
    ws.add_chart(pie_chart, "I19")

    # Cosmetic formatting
    for col, width in {
        "A": 24,
        "B": 16,
        "C": 16,
        "D": 4,
        "E": 28,
        "F": 16,
        "G": 34,
        "H": 4,
        "I": 4,
    }.items():
        ws.column_dimensions[col].width = width

    for col in ["J", "K", "L", "M", "N", "O"]:
        ws.column_dimensions[col].hidden = True

    for row in range(4, 30):
        ws.row_dimensions[row].height = 22


def _autosize_sheet(ws, max_width: int = 120) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        if ws.column_dimensions[column_letter].hidden:
            continue
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            value_length = len(str(cell.value))
            if value_length > max_length:
                max_length = value_length
        adjusted_width = min(max_length + 2, max_width)
        if adjusted_width > 0:
            ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 0, adjusted_width)


def build_dashboard() -> Path:
    report_text = DISCREPANCY_REPORT_PATH.read_text(encoding="utf-8")
    _, _, merged_balance = _load_balance_frames()
    financial = _financial_kpis(merged_balance)
    operational = _operational_kpis()

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_methodology = wb.create_sheet("Metodología y Diseño")
    ws_reconciliation = wb.create_sheet("Reconciliación")
    ws_dashboard = wb.create_sheet("Dashboard")

    _write_methodology_sheet(ws_methodology)
    _write_reconciliation_sheet(ws_reconciliation, report_text)
    _write_dashboard_sheet(ws_dashboard, financial, operational)

    # Make the sheets feel polished
    for ws in [ws_methodology, ws_reconciliation, ws_dashboard]:
        ws.sheet_view.showGridLines = False

    _autosize_sheet(ws_methodology)
    _autosize_sheet(ws_reconciliation)
    _autosize_sheet(ws_dashboard)

    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


def main() -> None:
    output_path = build_dashboard()
    print(f"Dashboard generated: {output_path}")


if __name__ == "__main__":
    main()
